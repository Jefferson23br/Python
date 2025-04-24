# Para criar novas abas na planilha automaticamente em Python de acordo com o preenchimento da planilha, você usará a biblioteca openpyxl.
# O processo envolve verificar o número de linhas ocupadas em uma planilha e criar novas abas com base nesse número. Aqui está o código para este fim:

#Importar Biblioteca
import openpyxl

# Carregar a planilha
workbook = openpyxl.load_workbook('Caminho Completo do local da Planilha')
sheet = workbook.active

# Definir o número máximo de linhas preenchidas para cada nova aba
max_rows_per_sheet = 100

# Contar o número de linhas preenchidas na planilha
num_rows = sheet.max_row

# Calcular o número de novas abas necessárias com base no número de linhas preenchidas
num_sheets = (num_rows // max_rows_per_sheet) + 1

# Criar novas abas
for i in range(num_sheets - 1):
    new_sheet = workbook.create_sheet(title=f'Sheet {i+2}')
    # Copiar os cabeçalhos da planilha original para as novas abas
    for j in range(1, sheet.max_column+1):
        new_sheet.cell(row=1, column=j).value = sheet.cell(row=1, column=j).value
    # Copiar as linhas correspondentes para as novas abas
    for j in range(max_rows_per_sheet):
        if sheet.cell(row=i*max_rows_per_sheet+j+2, column=1).value is not None:
            for k in range(1, sheet.max_column+1):
                new_sheet.cell(row=j+2, column=k).value = sheet.cell(row=i*max_rows_per_sheet+j+2, column=k).value

# Salvar a planilha
workbook.save('nome_do_arquivo.xlsx')

#Neste exemplo, o código cria novas abas com o nome "Planilha 2", "Planilha 3", etc.,dependendo do número de linhas atendidas na planilha original. 
#Cada nova aba tem um número máximo de linhas acomodadas definidas pela variável max_rows_per_sheet. As informações das novas abas são copiadas da planilha original, 
# incluindo cabeçalhos e linhas correspondentes.
# Finalmente, a planilha atualizada é salva em um arquivo